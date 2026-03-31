#!/usr/bin/env python3
"""
Extract actual database data from PostgreSQL dump and convert to Excel
"""
import os
import sys
import re
import struct
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

class PostgresDumpParser:
    """Parser for PostgreSQL binary dump files"""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.tables = defaultdict(list)  # table_name -> list of records
        self.table_schemas = {}  # table_name -> list of column names
        
    def parse(self):
        """Parse the dump file"""
        print("Attempting to parse PostgreSQL dump...")
        
        # Try different parsing methods
        if self._try_text_sections():
            print("✓ Successfully extracted data from text sections")
            return True
        
        print("⚠ Could not extract data using standard method")
        return False
    
    def _try_text_sections(self):
        """Try to extract readable text sections from the binary file"""
        try:
            with open(self.filepath, 'rb') as f:
                content = f.read()
            
            # Look for CREATE TABLE statements
            text_content = content.decode('utf-8', errors='ignore')
            
            # Extract CREATE TABLE statements
            self._extract_table_schemas(text_content)
            
            # Extract INSERT statements
            self._extract_insert_statements(text_content)
            
            return len(self.tables) > 0 or len(self.table_schemas) > 0
            
        except Exception as e:
            print(f"Error during parsing: {e}")
            return False
    
    def _extract_table_schemas(self, text_content):
        """Extract table schemas from CREATE TABLE statements"""
        # Pattern: CREATE TABLE tablename (...columns...)
        pattern = r'CREATE TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:public\.)?"?(\w+)"?\s*\((.*?)\);'
        matches = re.finditer(pattern, text_content, re.DOTALL | re.IGNORECASE)
        
        for match in matches:
            table_name = match.group(1)
            columns_def = match.group(2)
            
            columns = self._parse_columns(columns_def)
            if columns:
                self.table_schemas[table_name] = columns
                print(f"  Found table: {table_name} ({len(columns)} columns)")
    
    def _parse_columns(self, columns_def):
        """Parse column names from CREATE TABLE definition"""
        columns = []
        lines = columns_def.split(',')
        
        for line in lines:
            line = line.strip()
            if not line or any(x in line.upper() for x in ['CONSTRAINT', 'PRIMARY KEY', 'FOREIGN KEY', 'INDEX', 'UNIQUE']):
                continue
            
            # Extract column name (first word)
            parts = line.split(None, 1)
            if parts:
                col_name = parts[0].strip('"`')
                columns.append(col_name)
        
        return columns
    
    def _extract_insert_statements(self, text_content):
        """Extract data from INSERT statements"""
        # Pattern: INSERT INTO tablename (columns) VALUES (values)
        pattern = r'INSERT INTO\s+(?:public\.)?["\']?(\w+)["\']?\s*(?:\((.*?)\))?\s+VALUES\s*(.*?)(?:;|$)'
        matches = re.finditer(pattern, text_content, re.DOTALL | re.IGNORECASE)
        
        for match in matches:
            table_name = match.group(1)
            columns = match.group(2)
            values = match.group(3)
            
            if table_name not in self.table_schemas and columns:
                # Parse columns from INSERT statement
                col_list = [c.strip().strip('"`') for c in columns.split(',')]
                self.table_schemas[table_name] = col_list
            
            # Parse values
            record = self._parse_values(values)
            if record:
                self.tables[table_name].append(record)
        
        # Print summary
        for table_name, records in self.tables.items():
            if records:
                print(f"  Extracted {table_name}: {len(records)} rows")
    
    def _parse_values(self, values_str):
        """Parse VALUES clause to extract data"""
        values_str = values_str.strip()
        if values_str.startswith('('):
            values_str = values_str[1:]
        if values_str.endswith(')'):
            values_str = values_str[:-1]
        
        # Simple parser for values
        result = []
        current = ''
        in_string = False
        escape = False
        
        for char in values_str:
            if escape:
                current += char
                escape = False
            elif char == '\\':
                escape = True
            elif char == "'":
                in_string = not in_string
                current += char
            elif char == ',' and not in_string:
                result.append(self._clean_value(current))
                current = ''
            else:
                current += char
        
        if current:
            result.append(self._clean_value(current))
        
        return result if result else None
    
    def _clean_value(self, val):
        """Clean and format extracted values"""
        val = val.strip()
        # Remove quotes
        if (val.startswith("'") and val.endswith("'")) or (val.startswith('"') and val.endswith('"')):
            val = val[1:-1]
        # Handle escapes
        val = val.replace("\\'", "'").replace('\\"', '"').replace('\\\\', '\\')
        return val if val.upper() != 'NULL' else ''

def create_data_excel(tables_data, table_schemas, output_filepath):
    """Create Excel file with actual data"""
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet
    summary_sheet = wb.create_sheet('Summary', 0)
    summary_sheet['A1'] = 'Database Data Overview'
    summary_sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    summary_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    summary_sheet.merge_cells('A1:C1')
    
    summary_sheet['A3'] = 'Table Name'
    summary_sheet['B3'] = 'Rows'
    summary_sheet['C3'] = 'Columns'
    
    for col in ['A', 'B', 'C']:
        summary_sheet[f'{col}3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        summary_sheet[f'{col}3'].font = Font(bold=True)
    
    row = 4
    total_rows = 0
    for table_name in sorted(tables_data.keys()):
        rows_count = len(tables_data[table_name])
        cols_count = len(table_schemas.get(table_name, []))
        
        summary_sheet[f'A{row}'] = table_name
        summary_sheet[f'B{row}'] = rows_count
        summary_sheet[f'C{row}'] = cols_count
        total_rows += rows_count
        row += 1
    
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 12
    summary_sheet.column_dimensions['C'].width = 12
    
    # Create sheets for tables with data
    sheet_count = 0
    for table_name in sorted(tables_data.keys()):
        records = tables_data[table_name]
        columns = table_schemas.get(table_name, [])
        
        if not records or not columns:
            continue
        
        # Ensure sheet name is valid (max 31 chars)
        sheet_name = table_name[:31]
        sheet = wb.create_sheet(sheet_name)
        sheet_count += 1
        
        # Header
        sheet['A1'] = f'Table: {table_name}'
        sheet['A1'].font = Font(size=12, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sheet.merge_cells(f'A1:{chr(64 + len(columns))}1')
        
        # Column headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = col_name
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Add data
        for row_idx, record in enumerate(records, 4):
            for col_idx, value in enumerate(record, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for record in records:
                if col_idx <= len(record):
                    max_length = max(max_length, len(str(record[col_idx - 1])))
            
            col_letter = chr(64 + col_idx)
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        sheet.row_dimensions[3].height = 25
    
    wb.save(output_filepath)
    print(f"\n✓ Excel file created: {output_filepath}")
    print(f"✓ Tables with data: {sheet_count}")
    print(f"✓ Total rows exported: {total_rows}")

def main():
    sql_file = 'convergeai_conference_db.sql'
    current_dir = Path(__file__).parent
    sql_filepath = current_dir / sql_file
    
    if not sql_filepath.exists():
        print(f"Error: {sql_file} not found in {current_dir}")
        sys.exit(1)
    
    print(f"📖 Parsing PostgreSQL dump: {sql_filepath}")
    print(f"📊 File size: {sql_filepath.stat().st_size / 1024:.2f} KB\n")
    
    parser = PostgresDumpParser(str(sql_filepath))
    success = parser.parse()
    
    if not success or (not parser.tables and not parser.table_schemas):
        print("\n❌ Could not extract data from the binary dump.")
        print("\nNote: This appears to be a PostgreSQL binary dump (PGDUMP format).")
        print("To extract actual data, you may need to:")
        print("1. Restore the dump to a PostgreSQL database:")
        print(f"   pg_restore -d database_name {sql_file}")
        print("\n2. Or export as text SQL using:")
        print(f"   pg_restore -f output.sql {sql_file}")
        print("\nAlternatively, if you have the .sql text file, the script will extract data from INSERT statements.")
        sys.exit(1)
    
    output_file = current_dir / 'database_with_data.xlsx'
    print(f"\n📝 Creating Excel file with data...")
    create_data_excel(parser.tables, parser.table_schemas, str(output_file))
    print("\n✅ Conversion completed!")

if __name__ == '__main__':
    main()
