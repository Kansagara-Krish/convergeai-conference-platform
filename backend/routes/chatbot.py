# ============================================
# Chatbot Management Routes
# ============================================

from flask import Blueprint, request, jsonify, current_app
from datetime import datetime
from werkzeug.utils import secure_filename
from pathlib import Path
import csv
import json
import os
import uuid

from openpyxl import load_workbook

try:
    from models import db, Chatbot, Guest, Message, AdminNotification
    from routes.auth import token_required, admin_required
except ImportError:
    from backend.models import db, Chatbot, Guest, Message, AdminNotification
    from backend.routes.auth import token_required, admin_required

chatbot_bp = Blueprint('chatbot', __name__)

IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
GUEST_LIST_EXTENSIONS = {'csv', 'xlsx'}
GUEST_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
DEFAULT_SINGLE_PERSON_PROMPT = (
    'Generate a high-quality professional portrait image of the guest.\n\n'
    'Details:\n'
    '- Focus on one person only.\n'
    '- Center the person in the frame.\n'
    '- Use a given background image\n'
    '- Maintain realistic facial features.\n'
    '- Proper lighting and sharp focus.\n'
    '- Business or formal attire.\n'
    '- No extra people in the frame.\n'
    '- No distortion or overlapping elements.\n'
    '- Professional conference vibe.'
)
DEFAULT_MULTIPLE_PERSON_PROMPT = (
    'Generate a professional group image of multiple guests.\n\n'
    'Requirements:\n'
    '- Include all selected guests in one frame.\n'
    '- Arrange them naturally in a group.\n'
    '- Maintain correct proportions for each person.\n'
    '- Ensure no unnatural gaps between group members.\n'
    '- If people are close together, blend them naturally without visual separation.\n'
    '- Avoid cutting faces or overlapping distortions.\n'
    '- Use a conference or stage background.\n'
    '- Maintain uniform lighting and perspective.\n'
    '- Make the group appear cohesive and professionally composed.'
)


def to_bool(value, default=False):
    """Normalize common request boolean formats to real booleans."""
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in ('1', 'true', 'yes', 'on'):
            return True
        if normalized in ('0', 'false', 'no', 'off', ''):
            return False

    return default


def get_request_data():
    """Get request payload from JSON or multipart form data."""
    if request.form:
        return request.form.to_dict(), True

    data = request.get_json(silent=True)
    if isinstance(data, dict):
        return data, False

    return None, False


def get_file_extension(filename):
    """Return lowercase file extension without dot."""
    if not filename or '.' not in filename:
        return ''
    return filename.rsplit('.', 1)[1].lower()


def parse_end_date(value):
    """Parse optional end date; fallback to infinite when empty."""
    if value is None:
        return Chatbot.INFINITE_END_DATE

    value_str = str(value).strip()
    if not value_str:
        return Chatbot.INFINITE_END_DATE

    return datetime.fromisoformat(value_str).date()


def is_allowed_file(filename, allowed_extensions):
    """Validate file extension against allowed set."""
    return get_file_extension(filename) in allowed_extensions


def get_upload_directory(subdirectory):
    """Build and create upload directory path."""
    base_upload = current_app.config.get('UPLOAD_FOLDER', 'uploads')
    if not os.path.isabs(base_upload):
        base_upload = os.path.join(current_app.root_path, base_upload)

    final_dir = os.path.join(base_upload, subdirectory)
    os.makedirs(final_dir, exist_ok=True)
    return final_dir


def save_uploaded_file(file_obj, subdirectory):
    """Save uploaded file and return relative path."""
    upload_dir = get_upload_directory(subdirectory)
    original_name = secure_filename(file_obj.filename)
    unique_name = f"{uuid.uuid4().hex}_{original_name}"
    destination_path = os.path.join(upload_dir, unique_name)
    file_obj.save(destination_path)

    relative_path = Path('uploads') / subdirectory / unique_name
    return destination_path, str(relative_path).replace('\\', '/')


def normalize_guest_row(row):
    """Normalize guest columns from CSV/XLSX rows."""
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized_key = str(key).strip().lower().replace(' ', '_')
        normalized[normalized_key] = ('' if value is None else str(value).strip())
    return normalized


def parse_guest_list_file(file_path, extension):
    """Parse guest list file and return normalized guest entries."""
    guests = []

    if extension == 'csv':
        with open(file_path, newline='', encoding='utf-8-sig') as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                normalized = normalize_guest_row(row)
                if normalized.get('name'):
                    guests.append(normalized)
        return guests

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return guests

    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    for row_values in rows[1:]:
        row_dict = {
            headers[index]: row_values[index] if index < len(row_values) else ''
            for index in range(len(headers))
            if headers[index]
        }
        normalized = normalize_guest_row(row_dict)
        if normalized.get('name'):
            guests.append(normalized)

    return guests


def normalize_file_key(value):
    if value is None:
        return ''
    return secure_filename(str(value).strip()).lower()


def build_guest_image_lookup(image_files):
    lookup = {}
    for file_obj in image_files:
        filename = secure_filename(file_obj.filename or '')
        if not filename:
            continue
        normalized = filename.lower()
        stem = os.path.splitext(normalized)[0]
        lookup[normalized] = file_obj
        if stem and stem not in lookup:
            lookup[stem] = file_obj
    return lookup


def get_guest_image_reference(guest_row):
    return (
        guest_row.get('image_name')
        or guest_row.get('photo_name')
        or guest_row.get('image')
        or guest_row.get('photo')
        or ''
    )


def parse_json_list(raw_value):
    if not raw_value:
        return []
    try:
        parsed = json.loads(raw_value)
        return parsed if isinstance(parsed, list) else []
    except (TypeError, ValueError, json.JSONDecodeError):
        return []

# ============================================
# Create Chatbot
# ============================================

@chatbot_bp.route('', methods=['POST'])
@token_required
@admin_required
def create_chatbot(user):
    """Create a new chatbot"""

    data, is_form_data = get_request_data()
    if not isinstance(data, dict):
        return jsonify({'success': False, 'message': 'Invalid request body'}), 400

    # Validate required fields
    required_fields = ['name', 'event_name', 'start_date', 'single_person_prompt', 'multiple_person_prompt']
    if not all(str(data.get(field, '')).strip() for field in required_fields):
        return jsonify({'success': False, 'message': 'Missing required fields'}), 400

    system_prompt = str(data.get('system_prompt', '')).strip()
    single_person_prompt = str(data.get('single_person_prompt', '')).strip()
    multiple_person_prompt = str(data.get('multiple_person_prompt', '')).strip()

    if not system_prompt:
        system_prompt = single_person_prompt

    if not is_form_data:
        return jsonify({
            'success': False,
            'message': 'Use multipart/form-data with background_image file'
        }), 400

    background_image = request.files.get('background_image')
    guest_list = request.files.get('guest_list')
    guest_images = request.files.getlist('guest_images')
    manual_guests = parse_json_list(data.get('manual_guests'))

    has_excel_guest_list = bool(guest_list and guest_list.filename)
    has_manual_guests = len(manual_guests) > 0

    if not has_excel_guest_list and not has_manual_guests:
        return jsonify({
            'success': False,
            'message': 'Add at least one manual guest (name + image) or upload guest Excel file'
        }), 400

    if not background_image or not background_image.filename:
        return jsonify({'success': False, 'message': 'Background image is required'}), 400

    if not is_allowed_file(background_image.filename, IMAGE_EXTENSIONS):
        return jsonify({'success': False, 'message': 'Invalid background image type'}), 400

    if guest_list and guest_list.filename and not is_allowed_file(guest_list.filename, GUEST_LIST_EXTENSIONS):
        return jsonify({'success': False, 'message': 'Guest list must be .csv or .xlsx'}), 400

    for index, manual_guest in enumerate(manual_guests):
        name = str(manual_guest.get('name', '')).strip()
        photo_field = str(manual_guest.get('photo_file_field', '')).strip()

        if not name:
            return jsonify({
                'success': False,
                'message': f'Manual guest #{index + 1} requires a name'
            }), 400

        if not photo_field:
            return jsonify({
                'success': False,
                'message': f'Manual guest #{index + 1} requires an image'
            }), 400

        manual_photo_file = request.files.get(photo_field)
        if not manual_photo_file or not manual_photo_file.filename:
            return jsonify({
                'success': False,
                'message': f'Manual guest #{index + 1} image file is missing'
            }), 400

        if not is_allowed_file(manual_photo_file.filename, GUEST_IMAGE_EXTENSIONS):
            return jsonify({
                'success': False,
                'message': f'Invalid image type for manual guest #{index + 1}'
            }), 400

    for guest_image in guest_images:
        if guest_image and guest_image.filename and not is_allowed_file(guest_image.filename, GUEST_IMAGE_EXTENSIONS):
            return jsonify({'success': False, 'message': 'Guest images must be valid image files'}), 400

    saved_background_abs = None
    saved_guest_list_abs = None
    saved_guest_image_abs_paths = []
    created_successfully = False
    guest_rows = []
    guest_photo_lookup = {}
    
    try:
        saved_background_abs, background_image_path = save_uploaded_file(background_image, 'backgrounds')

        if guest_list and guest_list.filename:
            saved_guest_list_abs, _ = save_uploaded_file(guest_list, 'guest_lists')

            guest_rows = parse_guest_list_file(
                saved_guest_list_abs,
                get_file_extension(guest_list.filename)
            )

            if len(guest_rows) == 0:
                return jsonify({
                    'success': False,
                    'message': 'Guest list must contain at least one row with a name column'
                }), 400

        guest_image_lookup = build_guest_image_lookup(guest_images)
        for key, file_obj in guest_image_lookup.items():
            # Save each unique uploaded image only once (by full normalized filename keys)
            if '.' not in key:
                continue
            saved_abs, saved_relative = save_uploaded_file(file_obj, 'guests')
            saved_guest_image_abs_paths.append(saved_abs)
            guest_photo_lookup[key] = saved_relative
            stem = os.path.splitext(key)[0]
            if stem and stem not in guest_photo_lookup:
                guest_photo_lookup[stem] = saved_relative

        chatbot = Chatbot(
            name=data['name'],
            event_name=data['event_name'],
            description=data.get('description', ''),
            start_date=datetime.fromisoformat(data['start_date']).date(),
            end_date=parse_end_date(data.get('end_date')),
            system_prompt=system_prompt,
            single_person_prompt=single_person_prompt,
            multiple_person_prompt=multiple_person_prompt,
            background_image=background_image_path,
            public=to_bool(data.get('public'), True),
            active=to_bool(data.get('active'), True),
            created_by_id=user.id
        )

        db.session.add(chatbot)

        for guest_row in guest_rows:
            image_reference = normalize_file_key(get_guest_image_reference(guest_row))
            photo_path = guest_photo_lookup.get(image_reference) if image_reference else None
            guest = Guest(
                chatbot=chatbot,
                name=guest_row.get('name', ''),
                title=guest_row.get('title') or guest_row.get('designation') or guest_row.get('role') or '',
                description=guest_row.get('description') or guest_row.get('bio') or '',
                photo=photo_path,
                organization=guest_row.get('organization') or guest_row.get('company') or '',
                email=guest_row.get('email') or '',
                is_speaker=to_bool(guest_row.get('is_speaker') or guest_row.get('speaker'), False),
                is_moderator=to_bool(guest_row.get('is_moderator') or guest_row.get('moderator'), False)
            )
            db.session.add(guest)

        for index, manual_guest in enumerate(manual_guests):
            name = str(manual_guest.get('name', '')).strip()
            if not name:
                continue

            photo_field = str(manual_guest.get('photo_file_field', '')).strip()
            photo_path = None
            if photo_field:
                manual_photo_file = request.files.get(photo_field)
                if manual_photo_file and manual_photo_file.filename:
                    saved_abs, saved_relative = save_uploaded_file(manual_photo_file, 'guests')
                    saved_guest_image_abs_paths.append(saved_abs)
                    photo_path = saved_relative

            guest = Guest(
                chatbot=chatbot,
                name=name,
                photo=photo_path,
                title='',
                description='',
                organization='',
                email=''
            )
            db.session.add(guest)

        notification = AdminNotification(
            title='New chatbot created',
            message=f'"{chatbot.name}" for event "{chatbot.event_name}" was created.',
            entity_type='chatbot',
            entity_id=chatbot.id,
            is_read=False
        )
        db.session.add(notification)

        db.session.commit()
        created_successfully = True
        
        return jsonify({
            'success': True,
            'message': 'Chatbot created successfully',
            'data': chatbot.to_dict(),
            'guests_imported': len(guest_rows),
            'manual_guests_added': len([g for g in manual_guests if str(g.get('name', '')).strip()])
        }), 201
    
    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Invalid date format: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error creating chatbot: {str(e)}'}), 500
    finally:
        if not created_successfully:
            for file_path in (saved_background_abs, saved_guest_list_abs, *saved_guest_image_abs_paths):
                if file_path and os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except OSError:
                        pass

# ============================================
# Update Chatbot
# ============================================

@chatbot_bp.route('/<int:chatbot_id>', methods=['PUT'])
@token_required
@admin_required
def update_chatbot(user, chatbot_id):
    """Update chatbot"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    data, is_form_data = get_request_data()
    manual_guests = parse_json_list(data.get('manual_guests'))
    deleted_guest_ids = parse_json_list(data.get('deleted_guest_ids'))


    if not isinstance(data, dict):
        return jsonify({'success': False, 'message': 'Invalid request body'}), 400
    
    # Update fields
    if 'name' in data:
        chatbot.name = data['name']
    if 'event_name' in data:
        chatbot.event_name = data['event_name']
    if 'description' in data:
        chatbot.description = data['description']
    if 'system_prompt' in data:
        if not str(data['system_prompt']).strip():
            return jsonify({'success': False, 'message': 'System prompt cannot be empty'}), 400
        chatbot.system_prompt = data['system_prompt']
    if 'single_person_prompt' in data:
        single_person_prompt = str(data['single_person_prompt']).strip()
        if not single_person_prompt:
            return jsonify({'success': False, 'message': 'Single person prompt cannot be empty'}), 400
        chatbot.single_person_prompt = single_person_prompt
    if 'multiple_person_prompt' in data:
        multiple_person_prompt = str(data['multiple_person_prompt']).strip()
        if not multiple_person_prompt:
            return jsonify({'success': False, 'message': 'Multiple person prompt cannot be empty'}), 400
        chatbot.multiple_person_prompt = multiple_person_prompt
    if 'public' in data:
        chatbot.public = to_bool(data['public'], chatbot.public)
    if 'active' in data:
        chatbot.active = to_bool(data['active'], chatbot.active)

    if is_form_data:
        background_image = request.files.get('background_image')
        guest_list = request.files.get('guest_list')
        guest_images = request.files.getlist('guest_images')

        if background_image and background_image.filename:
            if not is_allowed_file(background_image.filename, IMAGE_EXTENSIONS):
                return jsonify({'success': False, 'message': 'Invalid background image type'}), 400

            _, background_image_path = save_uploaded_file(background_image, 'backgrounds')
            chatbot.background_image = background_image_path

        if guest_list and guest_list.filename and not is_allowed_file(guest_list.filename, GUEST_LIST_EXTENSIONS):
            return jsonify({'success': False, 'message': 'Guest list must be .csv or .xlsx'}), 400

        for guest_image in guest_images:
            if guest_image and guest_image.filename and not is_allowed_file(guest_image.filename, GUEST_IMAGE_EXTENSIONS):
                return jsonify({'success': False, 'message': 'Guest images must be valid image files'}), 400

        guest_photo_lookup = {}
        guest_image_lookup = build_guest_image_lookup(guest_images)
        for key, file_obj in guest_image_lookup.items():
            if '.' not in key:
                continue
            _, saved_relative = save_uploaded_file(file_obj, 'guests')
            guest_photo_lookup[key] = saved_relative
            stem = os.path.splitext(key)[0]
            if stem and stem not in guest_photo_lookup:
                guest_photo_lookup[stem] = saved_relative

        if guest_list and guest_list.filename:
            saved_guest_list_abs, _ = save_uploaded_file(guest_list, 'guest_lists')
            imported_rows = parse_guest_list_file(saved_guest_list_abs, get_file_extension(guest_list.filename))

            for guest_row in imported_rows:
                name = str(guest_row.get('name', '')).strip()
                if not name:
                    continue
                image_reference = normalize_file_key(get_guest_image_reference(guest_row))
                photo_path = guest_photo_lookup.get(image_reference) if image_reference else None
                db.session.add(Guest(
                    chatbot=chatbot,
                    name=name,
                    title=guest_row.get('title') or guest_row.get('designation') or guest_row.get('role') or '',
                    description=guest_row.get('description') or guest_row.get('bio') or '',
                    photo=photo_path,
                    organization=guest_row.get('organization') or guest_row.get('company') or '',
                    email=guest_row.get('email') or '',
                    is_speaker=to_bool(guest_row.get('is_speaker') or guest_row.get('speaker'), False),
                    is_moderator=to_bool(guest_row.get('is_moderator') or guest_row.get('moderator'), False)
                ))

        for raw_id in deleted_guest_ids:
            try:
                guest_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            existing_guest = Guest.query.filter_by(id=guest_id, chatbot_id=chatbot.id).first()
            if existing_guest:
                db.session.delete(existing_guest)

        for index, manual_guest in enumerate(manual_guests):
            name = str(manual_guest.get('name', '')).strip()
            if not name:
                continue

            photo_field = str(manual_guest.get('photo_file_field', '')).strip()
            photo_path = None
            if photo_field:
                manual_photo_file = request.files.get(photo_field)
                if manual_photo_file and manual_photo_file.filename:
                    if not is_allowed_file(manual_photo_file.filename, GUEST_IMAGE_EXTENSIONS):
                        return jsonify({'success': False, 'message': f'Invalid image type for manual guest #{index + 1}'}), 400
                    _, photo_path = save_uploaded_file(manual_photo_file, 'guests')

            guest_id = manual_guest.get('id')
            existing_guest = None
            if guest_id is not None:
                try:
                    guest_id = int(guest_id)
                    existing_guest = Guest.query.filter_by(id=guest_id, chatbot_id=chatbot.id).first()
                except (TypeError, ValueError):
                    existing_guest = None

            if existing_guest:
                existing_guest.name = name
                if photo_path:
                    existing_guest.photo = photo_path
            else:
                db.session.add(Guest(
                    chatbot=chatbot,
                    name=name,
                    photo=photo_path,
                    title='',
                    description='',
                    organization='',
                    email=''
                ))
        
        # Handle image clearing
        if data.get('clear_background_image') == '1':
            chatbot.background_image = None
    
    # Handle image clearing from JSON data
    if 'clear_background_image' in data and data.get('clear_background_image') == '1':
        chatbot.background_image = None
    
    if 'start_date' in data:
        chatbot.start_date = datetime.fromisoformat(data['start_date']).date()
    if 'end_date' in data:
        chatbot.end_date = parse_end_date(data['end_date'])
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Chatbot updated successfully',
        'data': chatbot.to_dict()
    }), 200

# ============================================
# Get Chatbot
# ============================================

@chatbot_bp.route('/<int:chatbot_id>', methods=['GET'])
@token_required
def get_chatbot(user, chatbot_id):
    """Get chatbot details"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    # Check access (user can only see public chatbots or their own)
    if not chatbot.public and chatbot.created_by_id != user.id and user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    chatbot_data = chatbot.to_dict()
    chatbot_data['guests'] = [guest.to_dict() for guest in chatbot.guests]
    
    return jsonify({
        'success': True,
        'data': chatbot_data
    }), 200

# ============================================
# Delete Chatbot Settings
# ============================================

@chatbot_bp.route('/<int:chatbot_id>/settings', methods=['GET'])
@token_required
@admin_required
def get_chatbot_settings(user, chatbot_id):
    """Get chatbot settings"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    return jsonify({
        'success': True,
        'data': {
            'id': chatbot.id,
            'name': chatbot.name,
            'system_prompt': chatbot.system_prompt,
            'single_person_prompt': chatbot.single_person_prompt or DEFAULT_SINGLE_PERSON_PROMPT,
            'multiple_person_prompt': chatbot.multiple_person_prompt or DEFAULT_MULTIPLE_PERSON_PROMPT,
            'public': chatbot.public,
            'active': chatbot.active
        }
    }), 200

# ============================================
# Statistics
# ============================================

@chatbot_bp.route('/<int:chatbot_id>/stats', methods=['GET'])
@token_required
def get_chatbot_stats(user, chatbot_id):
    """Get chatbot statistics"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    stats = {
        'total_messages': Message.query.filter_by(chatbot_id=chatbot_id).count(),
        'total_guests': len(chatbot.guests),
        'total_participants': len(chatbot.participants),
        'user_messages': Message.query.filter(
            Message.chatbot_id == chatbot_id,
            Message.is_user_message == True
        ).count(),
        'bot_messages': Message.query.filter(
            Message.chatbot_id == chatbot_id,
            Message.is_user_message == False
        ).count()
    }
    
    return jsonify({
        'success': True,
        'data': stats
    }), 200
