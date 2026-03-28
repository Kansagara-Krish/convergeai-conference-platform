# ============================================
# Admin Routes
# ============================================

from flask import Blueprint, request, jsonify, current_app
from datetime import datetime, timedelta
from openpyxl import load_workbook
from io import BytesIO
from sqlalchemy import or_, func, case, extract
from werkzeug.utils import secure_filename
import re
import smtplib
import ssl
from email.message import EmailMessage
import threading
import os
import uuid

try:
    from models import db, User, Chatbot, Guest, Message, SessionToken, ChatbotParticipant
    from routes.auth import token_required, admin_required
    from services.email_templates import build_user_credentials_email
except ImportError:
    from backend.models import db, User, Chatbot, Guest, Message, SessionToken, ChatbotParticipant
    from backend.routes.auth import token_required, admin_required
    from backend.services.email_templates import build_user_credentials_email

admin_bp = Blueprint('admin', __name__)
EMAIL_REGEX = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$')
GUEST_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
VALID_USER_ROLES = {'admin', 'user', 'speaker', 'volunteer'}
INDIA_WHATSAPP_REGEX = re.compile(r'^(?:\+91)?[6-9]\d{9}$')


def _to_bool(value, default=False):
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in ('1', 'true', 'yes', 'on'):
            return True
        if normalized in ('0', 'false', 'no', 'off', ''):
            return False

    return default


def _is_allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def _sanitize_image_stem(value):
    stem = secure_filename(str(value or '').strip())
    if not stem:
        return ''
    if '.' in stem:
        stem = stem.rsplit('.', 1)[0]
    return stem[:120]

def _build_unique_filename(upload_root, stem, extension):
    safe_stem = _sanitize_image_stem(stem) or uuid.uuid4().hex
    candidate = f"{safe_stem}.{extension}"
    counter = 1
    while os.path.exists(os.path.join(upload_root, candidate)):
        candidate = f"{safe_stem}_{counter}.{extension}"
        counter += 1
    return candidate


def _save_guest_image(file_obj, desired_name=None):
    if not file_obj or not file_obj.filename:
        return None

    filename = secure_filename(file_obj.filename)
    if not filename:
        return None

    if not _is_allowed_file(filename, GUEST_IMAGE_EXTENSIONS):
        return 'invalid-type'

    extension = filename.rsplit('.', 1)[1].lower()

    upload_root = os.path.join(current_app.root_path, 'uploads', 'guests')
    os.makedirs(upload_root, exist_ok=True)

    unique_filename = _build_unique_filename(upload_root, desired_name, extension)

    absolute_path = os.path.join(upload_root, unique_filename)
    file_obj.save(absolute_path)
    return f"uploads/guests/{unique_filename}"


def _rename_guest_image(existing_photo_path, desired_name):
    if not existing_photo_path:
        return None

    current_relative = str(existing_photo_path).replace('\\', '/')
    current_filename = current_relative.split('/')[-1]
    if '.' not in current_filename:
        return None

    extension = current_filename.rsplit('.', 1)[1].lower()
    upload_root = os.path.join(current_app.root_path, 'uploads', 'guests')
    os.makedirs(upload_root, exist_ok=True)

    current_abs = os.path.join(current_app.root_path, current_relative.replace('/', os.sep))
    if not os.path.exists(current_abs):
        return None

    new_filename = _build_unique_filename(upload_root, desired_name, extension)
    new_abs = os.path.join(upload_root, new_filename)

    if os.path.normcase(current_abs) == os.path.normcase(new_abs):
        return current_relative

    os.rename(current_abs, new_abs)
    return f"uploads/guests/{new_filename}"


def _normalize_upload_relative_path(raw_path):
    if not raw_path:
        return None

    normalized = str(raw_path).strip().replace('\\', '/').lstrip('/')
    if not normalized:
        return None

    if '/uploads/' in normalized:
        normalized = normalized.split('/uploads/', 1)[1]
        normalized = f"uploads/{normalized}"
    elif not normalized.startswith('uploads/'):
        if normalized.startswith(('backgrounds/', 'guests/', 'guest_lists/')):
            normalized = f"uploads/{normalized}"
        else:
            return None

    return normalized


def _resolve_upload_absolute_path(raw_path):
    relative_path = _normalize_upload_relative_path(raw_path)
    if not relative_path:
        return None

    uploads_root = os.path.abspath(os.path.join(current_app.root_path, 'uploads'))
    absolute_path = os.path.abspath(
        os.path.join(current_app.root_path, relative_path.replace('/', os.sep))
    )

    if absolute_path != uploads_root and not absolute_path.startswith(uploads_root + os.sep):
        return None

    return absolute_path


def _delete_uploaded_file(raw_path):
    absolute_path = _resolve_upload_absolute_path(raw_path)
    if not absolute_path:
        return
    if not os.path.exists(absolute_path):
        return
    try:
        os.remove(absolute_path)
    except OSError:
        pass


def is_valid_email(email):
    return bool(email and EMAIL_REGEX.match(email))


def _stringify_excel_cell(value):
    if value is None:
        return ''

    if isinstance(value, float):
        # Excel often returns phone numbers as floats like 9876543210.0.
        if value.is_integer():
            return str(int(value))
        return format(value, 'f').rstrip('0').rstrip('.')

    return str(value).strip()


def _normalize_excel_header_key(value):
    key = _stringify_excel_cell(value).lower()
    if not key:
        return ''
    return re.sub(r'[\s\-]+', '_', key)


def normalize_indian_whatsapp_number(raw_value):
    raw = _stringify_excel_cell(raw_value)
    if not raw:
        return None

    compact = re.sub(r'[^\d+]', '', raw)
    if compact.startswith('+91'):
        candidate = compact
    else:
        digits_only = re.sub(r'\D', '', compact)
        if digits_only.startswith('91') and len(digits_only) == 12:
            candidate = f'+{digits_only}'
        elif len(digits_only) == 10:
            candidate = f'+91{digits_only}'
        else:
            candidate = compact

    if not INDIA_WHATSAPP_REGEX.fullmatch(candidate):
        return None

    normalized_digits = re.sub(r'\D', '', candidate)
    if normalized_digits.startswith('91') and len(normalized_digits) == 12:
        return f'+{normalized_digits}'
    return None


def _make_unique_email(email):
    normalized_email = str(email or '').strip().lower()
    if not normalized_email:
        return normalized_email, False

    if User.query.filter_by(email=normalized_email).first() is None:
        return normalized_email, False

    local_part, separator, domain_part = normalized_email.partition('@')
    if not separator or not domain_part:
        return normalized_email, False

    safe_local = re.sub(r'[^a-z0-9._+-]', '', local_part) or 'user'
    suffix = 1

    while True:
        candidate = f"{safe_local}+{suffix}@{domain_part}"
        if User.query.filter_by(email=candidate).first() is None:
            return candidate, True
        suffix += 1


def _email_key(raw_email):
    return str(raw_email or '').strip().lower()


def _email_exists(raw_email):
    email_key = _email_key(raw_email)
    if not email_key:
        return False

    return db.session.query(User.id).filter(func.lower(User.email) == email_key).first() is not None


def _existing_email_keys(email_keys):
    normalized_keys = {str(value).strip().lower() for value in (email_keys or set()) if str(value).strip()}
    if not normalized_keys:
        return set()

    rows = db.session.query(User.email).filter(func.lower(User.email).in_(list(normalized_keys))).all()
    return {
        str(row[0]).strip().lower()
        for row in rows
        if row and row[0] is not None and str(row[0]).strip()
    }


def _make_unique_email_with_reserved(email, reserved_email_keys=None):
    reserved = reserved_email_keys or set()
    normalized_email = _email_key(email)
    if not normalized_email:
        return normalized_email, False

    if (
        normalized_email not in reserved
        and User.query.filter(func.lower(User.email) == normalized_email).first() is None
    ):
        return normalized_email, False

    local_part, separator, domain_part = normalized_email.partition('@')
    if not separator or not domain_part:
        return normalized_email, False

    safe_local = re.sub(r'[^a-z0-9._+-]', '', local_part) or 'user'
    suffix = 1

    while True:
        candidate = f"{safe_local}+{suffix}@{domain_part}"
        if candidate in reserved:
            suffix += 1
            continue
        if User.query.filter(func.lower(User.email) == candidate).first() is None:
            return candidate, True
        suffix += 1


def _username_key(raw_username):
    return str(raw_username or '').strip().lower()


def _username_exists(raw_username):
    username_key = _username_key(raw_username)
    if not username_key:
        return False

    return db.session.query(User.id).filter(func.lower(User.username) == username_key).first() is not None


def _existing_username_keys(username_keys):
    normalized_keys = {str(value).strip().lower() for value in (username_keys or set()) if str(value).strip()}
    if not normalized_keys:
        return set()

    rows = db.session.query(User.username).filter(func.lower(User.username).in_(list(normalized_keys))).all()
    return {
        str(row[0]).strip().lower()
        for row in rows
        if row and row[0] is not None and str(row[0]).strip()
    }


def send_user_credentials_email(recipient_email, name, role, username, password, allowed_events=None, allowed_chatbots=None):
    """Send credentials email synchronously"""
    mail_server = current_app.config.get('MAIL_SERVER')
    mail_port = current_app.config.get('MAIL_PORT', 587)
    mail_username = current_app.config.get('MAIL_USERNAME')
    mail_password = current_app.config.get('MAIL_PASSWORD')
    mail_sender = current_app.config.get('MAIL_DEFAULT_SENDER') or mail_username or 'noreply@localhost'
    use_tls = current_app.config.get('MAIL_USE_TLS', True)
    use_ssl = current_app.config.get('MAIL_USE_SSL', False)

    if not mail_server:
        return False, 'MAIL_SERVER is not configured'

    template = build_user_credentials_email(
        role=role,
        username=username,
        email=recipient_email,
        password=password,
        allowed_events=allowed_events,
        allowed_chatbots=allowed_chatbots,
    )

    msg = EmailMessage()
    msg['Subject'] = template['subject']
    msg['From'] = mail_sender
    msg['To'] = recipient_email
    msg.set_content(template['body'])

    try:
        if use_ssl:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(mail_server, mail_port, context=context, timeout=20) as server:
                if mail_username:
                    server.login(mail_username, mail_password or '')
                server.send_message(msg)
        else:
            with smtplib.SMTP(mail_server, mail_port, timeout=20) as server:
                if use_tls:
                    context = ssl.create_default_context()
                    server.starttls(context=context)
                if mail_username:
                    server.login(mail_username, mail_password or '')
                server.send_message(msg)
        return True, None
    except Exception as exc:
        return False, str(exc)


def send_email_background(app, recipient_email, name, role, username, password, allowed_events=None, allowed_chatbots=None):
    """Send email in background thread with app context"""
    with app.app_context():
        try:
            send_user_credentials_email(
                recipient_email,
                name,
                role,
                username,
                password,
                allowed_events,
                allowed_chatbots,
            )
        except Exception as e:
            # Log error but don't crash the thread
            print(f"Background email failed for {recipient_email}: {str(e)}")


def send_bulk_emails_background(app, credentials_list):
    """Send multiple emails in background thread with app context"""
    with app.app_context():
        for cred in credentials_list:
            try:
                send_user_credentials_email(
                    recipient_email=cred['email'],
                    name=cred['name'],
                    role=cred['role'],
                    username=cred.get('username') or '',
                    password=cred['password'],
                    allowed_events=cred.get('allowed_events'),
                    allowed_chatbots=cred.get('allowed_chatbots')
                )
            except Exception as e:
                print(f"Background email failed for {cred['email']}: {str(e)}")


def _parse_year_filter():
    raw_year = request.args.get('year', type=int)
    if raw_year is None:
        return None

    if raw_year < 2000 or raw_year > 2100:
        return 'invalid'

    return raw_year


def _parse_registration_year(raw_value):
    if raw_value is None:
        return None

    try:
        parsed = int(raw_value)
    except (TypeError, ValueError):
        return 'invalid'

    if parsed < 2000 or parsed > 2100:
        return 'invalid'

    return parsed


def _empty_image_stats(chatbot_id):
    return {
        'chatbot_id': int(chatbot_id),
        'total_images_generated': 0,
        'user_generated_count': 0,
        'volunteer_generated_count': 0,
    }


def _build_chatbot_image_stats_map(chatbot_ids=None, year=None):
    normalized_role = func.lower(func.coalesce(User.role, ''))
    is_user_role = normalized_role == 'user'
    is_volunteer_role = normalized_role == 'volunteer'

    stats_query = db.session.query(
        Message.chatbot_id.label('chatbot_id'),
        func.coalesce(
            func.sum(
                case(
                    (or_(is_user_role, is_volunteer_role), 1),
                    else_=0,
                )
            ),
            0,
        ).label('total_images_generated'),
        func.coalesce(
            func.sum(
                case(
                    (is_user_role, 1),
                    else_=0,
                )
            ),
            0,
        ).label('user_generated_count'),
        func.coalesce(
            func.sum(
                case(
                    (is_volunteer_role, 1),
                    else_=0,
                )
            ),
            0,
        ).label('volunteer_generated_count'),
    ).outerjoin(User, User.id == Message.user_id).filter(
        Message.is_user_message.is_(False),
        func.lower(Message.message_type) == 'image',
        Message.image_url.isnot(None),
        func.length(func.trim(Message.image_url)) > 0,
    )

    if chatbot_ids:
        stats_query = stats_query.filter(Message.chatbot_id.in_(chatbot_ids))

    if year:
        stats_query = stats_query.filter(extract('year', Message.created_at) == year)

    rows = stats_query.group_by(Message.chatbot_id).all()

    stats_map = {}
    for row in rows:
        chatbot_id = int(getattr(row, 'chatbot_id', 0) or 0)
        if chatbot_id <= 0:
            continue

        stats_map[chatbot_id] = {
            'chatbot_id': chatbot_id,
            'total_images_generated': int(getattr(row, 'total_images_generated', 0) or 0),
            'user_generated_count': int(getattr(row, 'user_generated_count', 0) or 0),
            'volunteer_generated_count': int(getattr(row, 'volunteer_generated_count', 0) or 0),
        }

    return stats_map

# ============================================
# Dashboard Statistics
# ============================================

@admin_bp.route('/dashboard/stats', methods=['GET'])
@token_required
@admin_required
def get_stats(user):
    """Get dashboard statistics"""
    now = datetime.now()
    first_day_of_month = datetime(now.year, now.month, 1)
    
    stats = {
        'total_chatbots': Chatbot.query.count(),
        'active_chatbots': Chatbot.query.filter_by(active=True).count(),
        'total_guests': Guest.query.count(),
        'total_users': User.query.count(),
        'active_users': User.query.filter_by(active=True).count(),
        'new_users_this_month': User.query.filter(User.created_at >= first_day_of_month).count(),
        'total_messages': Message.query.count(),
        'upcoming_events': Chatbot.query.filter(Chatbot.start_date > datetime.now().date()).count(),
    }
    
    return jsonify({
        'success': True,
        'data': stats
    }), 200


# ============================================
# User Management
# ============================================

@admin_bp.route('/users', methods=['GET'])
@token_required
@admin_required
def list_users(user):
    """List all users with pagination"""
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    role = request.args.get('role')
    active = request.args.get('active')
    year = request.args.get('year', type=int)
    chatbot_id = request.args.get('chatbot_id', type=int)
    search = (request.args.get('search') or '').strip()
    
    query = User.query.order_by(User.created_at.desc(), User.id.desc())
    
    if role:
        query = query.filter_by(role=role)
    
    if active is not None:
        active_bool = active.lower() == 'true'
        query = query.filter_by(active=active_bool)

    if year is not None:
        query = query.filter(extract('year', User.created_at) == year)

    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            or_(
                User.name.ilike(search_pattern),
                User.username.ilike(search_pattern)
            )
        )

    if chatbot_id:
        query = query.join(
            ChatbotParticipant,
            ChatbotParticipant.user_id == User.id
        ).filter(
            ChatbotParticipant.chatbot_id == chatbot_id
        ).distinct()
    
    paginated = query.paginate(page=page, per_page=per_page)
    
    # Include chatbot associations for each user
    users_data = []
    for u in paginated.items:
        user_dict = u.to_dict()
        user_dict['user_year'] = u.created_at.year if u.created_at else None
        # Get chatbots this user is associated with
        participants = ChatbotParticipant.query.filter_by(user_id=u.id).all()
        chatbot_data = []
        for participant in participants:
            chatbot = Chatbot.query.get(participant.chatbot_id)
            if chatbot:
                chatbot_data.append({
                    'id': chatbot.id,
                    'name': chatbot.name,
                    'event_name': chatbot.event_name
                })
        user_dict['chatbots'] = chatbot_data
        users_data.append(user_dict)
    
    return jsonify({
        'success': True,
        'data': users_data,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': page
    }), 200


@admin_bp.route('/users/years', methods=['GET'])
@token_required
@admin_required
def list_user_years(user):
    rows = db.session.query(
        extract('year', User.created_at).label('year')
    ).filter(
        User.created_at.isnot(None)
    ).group_by(
        extract('year', User.created_at)
    ).order_by(
        extract('year', User.created_at).desc()
    ).all()

    years = []
    for row in rows:
        year_value = getattr(row, 'year', None)
        if year_value is None:
            continue
        try:
            year_int = int(year_value)
        except (TypeError, ValueError):
            continue
        years.append(year_int)

    return jsonify({
        'success': True,
        'data': years,
    }), 200

@admin_bp.route('/users', methods=['POST'])
@token_required
@admin_required
def create_user(user):
    """Create a single user from admin panel"""

    data = request.get_json() or {}

    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    whatsapp_number_raw = data.get('whatsapp_number')
    role = (data.get('role') or 'user').strip().lower()
    active_raw = data.get('active', True)
    registration_year = _parse_registration_year(data.get('registration_year'))

    if registration_year == 'invalid':
        return jsonify({'success': False, 'message': 'registration_year must be between 2000 and 2100'}), 400

    if not email or not username or not password:
        return jsonify({'success': False, 'message': 'email, username, and password are required'}), 400

    # Name is optional in UI; default to username so DB nullable constraint is always satisfied.
    if not name:
        name = username

    normalized_whatsapp_number = normalize_indian_whatsapp_number(whatsapp_number_raw)
    if not normalized_whatsapp_number:
        return jsonify({'success': False, 'message': 'Valid Indian WhatsApp number is required'}), 400

    if not is_valid_email(email):
        return jsonify({'success': False, 'message': 'Invalid email format'}), 400

    if role not in VALID_USER_ROLES:
        return jsonify({'success': False, 'message': 'Invalid role'}), 400

    # Prevent regular admin users from creating other admin accounts
    if role == 'admin' and getattr(user, 'role', None) == 'admin':
        return jsonify({'success': False, 'message': 'You are not allowed to create admin users'}), 403

    effective_email, email_auto_adjusted = _make_unique_email(email)

    if _username_exists(username):
        return jsonify({'success': False, 'message': 'Username already exists'}), 409

    active = str(active_raw).strip().lower() in ['1', 'true', 'yes', 'y'] if isinstance(active_raw, str) else bool(active_raw)

    chatbot_ids = data.get('chatbot_ids', [])
    if chatbot_ids is None:
        chatbot_ids = []
    if not isinstance(chatbot_ids, list):
        return jsonify({'success': False, 'message': 'chatbot_ids must be a list'}), 400

    effective_user_year = registration_year or datetime.utcnow().year
    normalized_chatbot_ids = []
    for raw_chatbot_id in chatbot_ids:
        try:
            parsed_chatbot_id = int(raw_chatbot_id)
        except (TypeError, ValueError):
            continue
        if parsed_chatbot_id > 0 and parsed_chatbot_id not in normalized_chatbot_ids:
            normalized_chatbot_ids.append(parsed_chatbot_id)

    blocked_chatbots = []
    for chatbot_id in normalized_chatbot_ids:
        chatbot = Chatbot.query.get(chatbot_id)
        if not chatbot:
            continue

        chatbot_year = int(chatbot.start_date.year) if chatbot.start_date else None
        if (
            chatbot_year is not None
            and effective_user_year < chatbot_year
            and not bool(getattr(chatbot, 'allow_previous_year_users', False))
        ):
            blocked_chatbots.append(f"{chatbot.name} ({chatbot_year})")

    if blocked_chatbots:
        return jsonify({
            'success': False,
            'message': (
                'User year is older than selected event year and previous-year users are not allowed for: '
                + ', '.join(blocked_chatbots)
            )
        }), 400

    new_user = User(
        name=name,
        email=effective_email,
        username=username,
        whatsapp_number=normalized_whatsapp_number,
        role=role,
        active=active
    )
    if registration_year:
        new_user.created_at = datetime(registration_year, 1, 1)
    new_user.set_password(password)

    db.session.add(new_user)
    db.session.flush()

    # Associate user with chatbots if provided
    assigned_event_names = []
    assigned_chatbot_names = []
    if normalized_chatbot_ids:
        for chatbot_id in normalized_chatbot_ids:
            # Verify chatbot exists
            chatbot = Chatbot.query.get(chatbot_id)
            if chatbot:
                if chatbot.event_name and chatbot.event_name not in assigned_event_names:
                    assigned_event_names.append(chatbot.event_name)
                if chatbot.name and chatbot.name not in assigned_chatbot_names:
                    assigned_chatbot_names.append(chatbot.name)
                # Check if already a participant
                existing = ChatbotParticipant.query.filter_by(
                    chatbot_id=chatbot_id,
                    user_id=new_user.id
                ).first()
                
                if not existing:
                    participant = ChatbotParticipant(
                        chatbot_id=chatbot_id,
                        user_id=new_user.id
                    )
                    db.session.add(participant)

    db.session.commit()

    # Send email in background thread so admin doesn't wait
    app = current_app._get_current_object()
    thread = threading.Thread(
        target=send_email_background,
        args=(app, effective_email, name, role, username, password, assigned_event_names, assigned_chatbot_names)
    )
    thread.daemon = True
    thread.start()

    assigned_count = len(normalized_chatbot_ids) if normalized_chatbot_ids else 0
    db.session.commit()

    if assigned_count > 0:
        message = f'User created successfully and assigned to {assigned_count} event(s)! Credentials email is being sent.'
    else:
        message = 'User created successfully! Credentials email is being sent.'

    if email_auto_adjusted:
        message += f' Email already existed, so saved as {effective_email}.'

    return jsonify({
        'success': True,
        'message': message,
        'data': new_user.to_dict(),
        'chatbots_assigned': assigned_count,
        'email_auto_adjusted': email_auto_adjusted,
        'requested_email': email,
        'saved_email': effective_email,
        'registration_year': effective_user_year,
    }), 201

@admin_bp.route('/users/<int:user_id>', methods=['GET'])
@token_required
@admin_required
def get_user(user, user_id):
    """Get user details"""
    
    target_user = User.query.get(user_id)
    
    if not target_user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    user_data = target_user.to_dict()
    user_data['chatbots_created'] = len(target_user.chatbots)
    user_data['messages_sent'] = len(target_user.messages)
    
    return jsonify({
        'success': True,
        'data': user_data
    }), 200

@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@token_required
@admin_required
def update_user(user, user_id):
    """Update user details"""
    
    target_user = User.query.get(user_id)
    
    if not target_user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    data = request.get_json()
    
    # Allow updating specific fields
    if 'active' in data:
        target_user.active = data['active']
    if 'role' in data and data['role'] in VALID_USER_ROLES:
        # Prevent regular admin users from promoting others to admin
        if data['role'] == 'admin' and getattr(user, 'role', None) == 'admin':
            return jsonify({'success': False, 'message': 'You are not allowed to assign admin role'}), 403
        target_user.role = data['role']
    if 'name' in data:
        target_user.name = data['name']
    if 'whatsapp_number' in data:
        normalized_whatsapp_number = normalize_indian_whatsapp_number(data.get('whatsapp_number'))
        if not normalized_whatsapp_number:
            return jsonify({'success': False, 'message': 'Valid Indian WhatsApp number is required'}), 400
        target_user.whatsapp_number = normalized_whatsapp_number
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'User updated successfully',
        'data': target_user.to_dict()
    }), 200

@admin_bp.route('/users/<int:user_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_user(user, user_id):
    """Delete user"""
    
    if user_id == user.id:
        return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 400
    
    target_user = User.query.get(user_id)
    
    if not target_user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    # Delete all session tokens for this user first
    SessionToken.query.filter_by(user_id=user_id).delete()

    # Remove participant links so chatbot participant counts stay accurate
    ChatbotParticipant.query.filter_by(user_id=user_id).delete()
    
    # Now delete the user
    db.session.delete(target_user)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'User deleted successfully'
    }), 200


@admin_bp.route('/users/bulk-delete', methods=['POST', 'DELETE'])
@token_required
@admin_required
def bulk_delete_users(user):
    """Delete multiple users at once."""

    data = request.get_json(silent=True) or {}
    requested_ids = data.get('user_ids', [])

    if not isinstance(requested_ids, list) or not requested_ids:
        return jsonify({'success': False, 'message': 'user_ids list is required'}), 400

    normalized_ids = []
    for raw_id in requested_ids:
        try:
            parsed_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if parsed_id > 0 and parsed_id not in normalized_ids:
            normalized_ids.append(parsed_id)

    if not normalized_ids:
        return jsonify({'success': False, 'message': 'No valid user ids provided'}), 400

    self_skipped_ids = []
    if user.id in normalized_ids:
        self_skipped_ids.append(user.id)

    candidate_ids = [uid for uid in normalized_ids if uid != user.id]
    if not candidate_ids:
        return jsonify({
            'success': False,
            'message': 'Cannot delete your own account',
            'deleted_count': 0,
            'requested_count': len(normalized_ids),
            'skipped': {
                'self': self_skipped_ids,
                'not_found': []
            }
        }), 400

    existing_users = User.query.filter(User.id.in_(candidate_ids)).all()
    existing_ids = [u.id for u in existing_users]
    not_found_ids = [uid for uid in candidate_ids if uid not in existing_ids]

    if not existing_ids:
        return jsonify({
            'success': False,
            'message': 'No matching users found for deletion',
            'deleted_count': 0,
            'requested_count': len(normalized_ids),
            'skipped': {
                'self': self_skipped_ids,
                'not_found': not_found_ids,
            }
        }), 404

    SessionToken.query.filter(SessionToken.user_id.in_(existing_ids)).delete(synchronize_session=False)
    ChatbotParticipant.query.filter(ChatbotParticipant.user_id.in_(existing_ids)).delete(synchronize_session=False)

    # Use ORM row deletes so SQLAlchemy cascade rules remove dependent rows
    # (messages, conversations, guests, OTP records, etc.) before deleting users.
    for target_user in existing_users:
        db.session.delete(target_user)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'{len(existing_ids)} user(s) deleted successfully',
        'deleted_count': len(existing_ids),
        'requested_count': len(normalized_ids),
        'skipped': {
            'self': self_skipped_ids,
            'not_found': not_found_ids,
        }
    }), 200

@admin_bp.route('/users/bulk-activate', methods=['POST'])
@token_required
@admin_required
def bulk_activate_users(user):
    """Activate multiple users at once."""

    data = request.get_json(silent=True) or {}
    requested_ids = data.get('user_ids', [])

    if not isinstance(requested_ids, list) or not requested_ids:
        return jsonify({'success': False, 'message': 'user_ids list is required'}), 400

    normalized_ids = []
    for raw_id in requested_ids:
        try:
            parsed_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if parsed_id > 0 and parsed_id not in normalized_ids:
            normalized_ids.append(parsed_id)

    if not normalized_ids:
        return jsonify({'success': False, 'message': 'No valid user ids provided'}), 400

    existing_users = User.query.filter(User.id.in_(normalized_ids)).all()
    existing_ids = [u.id for u in existing_users]
    not_found_ids = [uid for uid in normalized_ids if uid not in existing_ids]

    if not existing_ids:
        return jsonify({
            'success': False,
            'message': 'No matching users found',
            'activated_count': 0,
            'requested_count': len(normalized_ids),
            'not_found': not_found_ids
        }), 404

    for target_user in existing_users:
        target_user.active = True

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'{len(existing_ids)} user(s) activated successfully',
        'activated_count': len(existing_ids),
        'requested_count': len(normalized_ids),
        'not_found': not_found_ids
    }), 200

@admin_bp.route('/users/bulk-deactivate', methods=['POST'])
@token_required
@admin_required
def bulk_deactivate_users(user):
    """Deactivate multiple users at once."""

    data = request.get_json(silent=True) or {}
    requested_ids = data.get('user_ids', [])

    if not isinstance(requested_ids, list) or not requested_ids:
        return jsonify({'success': False, 'message': 'user_ids list is required'}), 400

    normalized_ids = []
    for raw_id in requested_ids:
        try:
            parsed_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if parsed_id > 0 and parsed_id not in normalized_ids:
            normalized_ids.append(parsed_id)

    if not normalized_ids:
        return jsonify({'success': False, 'message': 'No valid user ids provided'}), 400

    self_skipped_ids = []
    if user.id in normalized_ids:
        self_skipped_ids.append(user.id)

    candidate_ids = [uid for uid in normalized_ids if uid != user.id]
    if not candidate_ids:
        return jsonify({
            'success': False,
            'message': 'Cannot deactivate your own account',
            'deactivated_count': 0,
            'requested_count': len(normalized_ids),
            'skipped': {
                'self': self_skipped_ids,
                'not_found': []
            }
        }), 400

    existing_users = User.query.filter(User.id.in_(candidate_ids)).all()
    existing_ids = [u.id for u in existing_users]
    not_found_ids = [uid for uid in candidate_ids if uid not in existing_ids]

    if not existing_ids:
        return jsonify({
            'success': False,
            'message': 'No matching users found for deactivation',
            'deactivated_count': 0,
            'requested_count': len(normalized_ids),
            'skipped': {
                'self': self_skipped_ids,
                'not_found': not_found_ids
            }
        }), 404

    for target_user in existing_users:
        target_user.active = False

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'{len(existing_ids)} user(s) deactivated successfully',
        'deactivated_count': len(existing_ids),
        'requested_count': len(normalized_ids),
        'skipped': {
            'self': self_skipped_ids,
            'not_found': not_found_ids
        }
    }), 200

# ============================================
# Chatbot Management
# ============================================

@admin_bp.route('/chatbots', methods=['GET'])
@token_required
@admin_required
def list_chatbots(user):
    """List all chatbots"""

    # Keep participant counts accurate by removing orphan participant rows
    # (can exist if users were deleted earlier without participant cleanup).
    ChatbotParticipant.query.filter(
        ~ChatbotParticipant.user_id.in_(db.session.query(User.id))
    ).delete(synchronize_session=False)
    db.session.commit()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    year = _parse_year_filter()

    if year == 'invalid':
        return jsonify({'success': False, 'message': 'year must be between 2000 and 2100'}), 400
    
    paginated = Chatbot.query.paginate(page=page, per_page=per_page)

    chatbot_items = list(paginated.items)
    chatbot_ids = [item.id for item in chatbot_items]
    image_stats_map = _build_chatbot_image_stats_map(chatbot_ids=chatbot_ids, year=year)

    data = []
    for chatbot in chatbot_items:
        chatbot_data = chatbot.to_dict()
        chatbot_data.update(image_stats_map.get(chatbot.id, _empty_image_stats(chatbot.id)))
        data.append(chatbot_data)
    
    return jsonify({
        'success': True,
        'data': data,
        'total': paginated.total,
        'pages': paginated.pages,
        'year': year,
    }), 200


@admin_bp.route('/chatbots/image-stats', methods=['GET'])
@token_required
@admin_required
def chatbot_image_stats(user):
    """Get chatbot-wise image generation counts with optional year filter."""

    year = _parse_year_filter()
    if year == 'invalid':
        return jsonify({'success': False, 'message': 'year must be between 2000 and 2100'}), 400

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    per_page = max(1, min(per_page, 200))

    paginated = Chatbot.query.order_by(Chatbot.created_at.desc(), Chatbot.id.desc()).paginate(page=page, per_page=per_page)
    chatbot_items = list(paginated.items)
    chatbot_ids = [item.id for item in chatbot_items]
    stats_map = _build_chatbot_image_stats_map(chatbot_ids=chatbot_ids, year=year)

    data = []
    for chatbot in chatbot_items:
        item = _empty_image_stats(chatbot.id)
        item.update(stats_map.get(chatbot.id, item))
        item['chatbot_name'] = chatbot.name
        item['event_name'] = chatbot.event_name
        data.append(item)

    return jsonify({
        'success': True,
        'data': data,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': paginated.page,
        'year': year,
    }), 200


@admin_bp.route('/chatbots/<int:chatbot_id>/image-count', methods=['GET'])
@token_required
@admin_required
def chatbot_image_count(user, chatbot_id):
    """Get image generation count for a specific chatbot."""

    chatbot = Chatbot.query.get(chatbot_id)
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404

    year = _parse_year_filter()
    if year == 'invalid':
        return jsonify({'success': False, 'message': 'year must be between 2000 and 2100'}), 400

    stats_map = _build_chatbot_image_stats_map(chatbot_ids=[chatbot_id], year=year)
    data = _empty_image_stats(chatbot_id)
    data.update(stats_map.get(chatbot_id, data))

    return jsonify({
        'success': True,
        'data': data,
        'year': year,
    }), 200


@admin_bp.route('/analytics', methods=['GET'])
@token_required
@admin_required
def analytics_overview(user):
    """Get chatbot image analytics with optional username filtering."""

    chatbot_id = request.args.get('chatbot_id', type=int)
    username = (request.args.get('username') or '').strip()

    if not chatbot_id:
        return jsonify({'message': 'chatbot_id is required'}), 400

    chatbot = Chatbot.query.get(chatbot_id)
    if not chatbot:
        return jsonify({'message': 'Chatbot not found'}), 404

    normalized_role = func.lower(func.coalesce(User.role, ''))
    is_user_role = normalized_role == 'user'
    is_volunteer_role = normalized_role == 'volunteer'

    filters = [
        Message.chatbot_id == chatbot_id,
        Message.is_user_message.is_(False),
        func.lower(Message.message_type) == 'image',
        Message.image_url.isnot(None),
        func.length(func.trim(Message.image_url)) > 0,
        or_(is_user_role, is_volunteer_role),
    ]

    if username:
        filters.append(User.username.ilike(f"%{username}%"))

    totals_row = db.session.query(
        func.coalesce(
            func.sum(
                case((is_user_role, 1), else_=0)
            ),
            0,
        ).label('user_images'),
        func.coalesce(
            func.sum(
                case((is_volunteer_role, 1), else_=0)
            ),
            0,
        ).label('volunteer_images'),
        func.coalesce(func.count(Message.id), 0).label('total_images'),
    ).join(User, User.id == Message.user_id).filter(*filters).first()

    timeline_rows = db.session.query(
        func.date(Message.created_at).label('date'),
        func.count(Message.id).label('count'),
    ).join(User, User.id == Message.user_id).filter(*filters).group_by(
        func.date(Message.created_at)
    ).order_by(
        func.date(Message.created_at).asc()
    ).all()

    breakdown_query = db.session.query(
        User.username.label('username'),
        func.count(Message.id).label('count'),
    ).join(User, User.id == Message.user_id).filter(*filters).group_by(
        User.id,
        User.username,
    ).order_by(
        func.count(Message.id).desc(),
        User.username.asc(),
    )

    if not username:
        breakdown_query = breakdown_query.limit(50)

    breakdown_rows = breakdown_query.all()

    return jsonify({
        'total_images': int(getattr(totals_row, 'total_images', 0) or 0),
        'user_images': int(getattr(totals_row, 'user_images', 0) or 0),
        'volunteer_images': int(getattr(totals_row, 'volunteer_images', 0) or 0),
        'timeline': [
            {
                'date': row.date.isoformat() if hasattr(row.date, 'isoformat') else str(row.date),
                'count': int(row.count or 0),
            }
            for row in timeline_rows
            if getattr(row, 'date', None) is not None
        ],
        'user_breakdown': [
            {
                'username': str(row.username or ''),
                'count': int(row.count or 0),
            }
            for row in breakdown_rows
        ],
    }), 200

@admin_bp.route('/chatbots/<int:chatbot_id>', methods=['GET'])
@token_required
@admin_required
def get_chatbot(user, chatbot_id):
    """Get chatbot details"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    chatbot_data = chatbot.to_dict()
    chatbot_data['guests'] = [guest.to_dict() for guest in chatbot.guests]
    
    return jsonify({
        'success': True,
        'data': chatbot_data
    }), 200

@admin_bp.route('/chatbots/<int:chatbot_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_chatbot(user, chatbot_id):
    """Delete chatbot and clean up related background/guest image files."""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404

    files_to_delete = []
    if chatbot.background_image:
        files_to_delete.append(chatbot.background_image)

    guest_photo_paths = [
        str(guest.photo).strip()
        for guest in (chatbot.guests or [])
        if getattr(guest, 'photo', None)
    ]
    
    db.session.delete(chatbot)
    db.session.commit()

    # Remove background file and guest image files that are no longer referenced.
    for file_path in set(files_to_delete):
        _delete_uploaded_file(file_path)

    for photo_path in set(guest_photo_paths):
        still_referenced = Guest.query.filter_by(photo=photo_path).first() is not None
        if not still_referenced:
            _delete_uploaded_file(photo_path)
    
    return jsonify({
        'success': True,
        'message': 'Chatbot deleted successfully'
    }), 200



# ============================================
# Guest Management
# ============================================

@admin_bp.route('/chatbots/<int:chatbot_id>/guests', methods=['GET'])
@token_required
@admin_required
def get_guests(user, chatbot_id):
    """Get chatbot guests"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    return jsonify({
        'success': True,
        'data': [guest.to_dict() for guest in chatbot.guests]
    }), 200

@admin_bp.route('/chatbots/<int:chatbot_id>/guests', methods=['POST'])
@token_required
@admin_required
def add_guest(user, chatbot_id):
    """Add guest to chatbot"""
    
    chatbot = Chatbot.query.get(chatbot_id)
    
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
    
    data = request.get_json()
    
    guest = Guest(
        chatbot_id=chatbot_id,
        name=data.get('name')
    )
    
    db.session.add(guest)
    db.session.flush()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Guest added successfully',
        'data': guest.to_dict()
    }), 201

@admin_bp.route('/guests', methods=['GET'])
@token_required
@admin_required
def list_all_guests(user):
    """List all guests across all chatbots"""
    
    # Eager load chatbot relationship to avoid N+1 queries
    guests = Guest.query.options(db.joinedload(Guest.chatbot)).all()
    
    return jsonify({
        'success': True,
        'data': [guest.to_dict() for guest in guests],
        'count': len(guests)
    }), 200


@admin_bp.route('/guests', methods=['POST'])
@token_required
@admin_required
def create_guest(user):
    """Create a new guest"""

    data = request.form.to_dict() if request.form else (request.get_json(silent=True) or {})

    if not isinstance(data, dict):
        return jsonify({'success': False, 'message': 'Invalid request payload'}), 400

    guest_name = str(data.get('name', '')).strip()
    chatbot_id = data.get('chatbot_id')

    if not guest_name:
        return jsonify({'success': False, 'message': 'Guest name is required'}), 400

    if not chatbot_id:
        return jsonify({'success': False, 'message': 'Please select chatbot/event'}), 400

    chatbot = Chatbot.query.get(chatbot_id)
    if not chatbot:
        return jsonify({'success': False, 'message': 'Chatbot not found'}), 404

    photo_path = None
    if 'photo' in request.files:
        photo_path = _save_guest_image(request.files.get('photo'), guest_name)
        if photo_path == 'invalid-type':
            return jsonify({'success': False, 'message': 'Invalid image type. Allowed: png, jpg, jpeg, gif, webp'}), 400
    
    guest = Guest(
        chatbot_id=chatbot.id,
        name=guest_name,
        photo=photo_path or None
    )
    
    db.session.add(guest)
    db.session.flush()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Guest created successfully',
        'data': guest.to_dict()
    }), 201



@admin_bp.route('/guests/<int:guest_id>', methods=['GET'])
@token_required
@admin_required
def get_guest(user, guest_id):
    """Get a single guest"""
    
    guest = Guest.query.options(db.joinedload(Guest.chatbot)).get(guest_id)
    
    if not guest:
        return jsonify({'success': False, 'message': 'Guest not found'}), 404
    
    return jsonify({
        'success': True,
        'data': guest.to_dict()
    }), 200


@admin_bp.route('/guests/<int:guest_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_guest(user, guest_id):
    """Delete a guest"""
    
    guest = Guest.query.get(guest_id)
    
    if not guest:
        return jsonify({'success': False, 'message': 'Guest not found'}), 404

    photo_path = guest.photo
    
    db.session.delete(guest)
    db.session.commit()

    if photo_path:
        _delete_uploaded_file(photo_path)
    
    return jsonify({
        'success': True,
        'message': 'Guest deleted successfully'
    }), 200


@admin_bp.route('/guests/<int:guest_id>', methods=['PUT'])
@token_required
@admin_required
def update_guest(user, guest_id):
    """Update a guest"""
    
    guest = Guest.query.get(guest_id)
    
    if not guest:
        return jsonify({'success': False, 'message': 'Guest not found'}), 404
    
    data = request.form.to_dict() if request.form else (request.get_json(silent=True) or {})

    if not isinstance(data, dict):
        return jsonify({'success': False, 'message': 'Invalid request payload'}), 400
    
    name_updated = False
    if 'name' in data and str(data.get('name', '')).strip():
        guest.name = str(data.get('name')).strip()
        name_updated = True
    if 'active' in data:
        guest.active = _to_bool(data.get('active'), guest.active)
    if 'chatbot_id' in data:
        # Validate chatbot exists if provided
        if data['chatbot_id']:
            chatbot = Chatbot.query.get(data['chatbot_id'])
            if not chatbot:
                return jsonify({'success': False, 'message': 'Chatbot not found'}), 404
        guest.chatbot_id = data['chatbot_id']

    auto_photo_name = guest.name

    if 'photo' in request.files:
        photo_path = _save_guest_image(request.files.get('photo'), auto_photo_name)
        if photo_path == 'invalid-type':
            return jsonify({'success': False, 'message': 'Invalid image type. Allowed: png, jpg, jpeg, gif, webp'}), 400
        if photo_path:
            guest.photo = photo_path
    elif name_updated and guest.photo:
        renamed_photo_path = _rename_guest_image(guest.photo, auto_photo_name)
        if renamed_photo_path:
            guest.photo = renamed_photo_path
    
    db.session.commit()
    
    # Reload with chatbot relationship
    db.session.refresh(guest)
    guest = Guest.query.options(db.joinedload(Guest.chatbot)).get(guest_id)
    
    return jsonify({
        'success': True,
        'message': 'Guest updated successfully',
        'data': guest.to_dict()
    }), 200

# ============================================
# Import Users
# ============================================

@admin_bp.route('/import/excel/preview', methods=['POST'])
@token_required
@admin_required
def preview_users_from_excel(user):
    """Preview users from Excel file without saving to DB"""

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400

    file = request.files['file']

    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': 'Only .xlsx files are allowed'}), 400

    try:
        workbook = load_workbook(filename=BytesIO(file.read()), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({'success': False, 'message': 'Excel file has no data rows'}), 400

        headers = [_normalize_excel_header_key(cell) if cell is not None else '' for cell in rows[0]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def get_value(row, keys, default=''):
            for key in keys:
                index = header_map.get(key)
                if index is not None and index < len(row):
                    value = row[index]
                    if value is not None:
                        return _stringify_excel_cell(value)
            return default

        total_rows = 0
        valid_rows = 0
        skipped_rows = 0
        missing_required_rows = 0
        duplicate_email_rows = 0
        existing_email_rows = 0
        duplicate_username_rows = 0
        existing_username_rows = 0
        invalid_email_rows = 0
        invalid_whatsapp_rows = 0
        missing_required_details = []
        duplicate_email_in_file_details = []
        existing_email_details = []
        invalid_email_details = []
        invalid_whatsapp_details = []
        duplicate_username_in_file_details = []
        existing_username_details = []
        max_preview_details = 20
        seen_email_keys = set()
        seen_username_keys = set()
        email_exists_cache = {}
        username_exists_cache = {}

        def email_exists_cached(email_key):
            if email_key not in email_exists_cache:
                email_exists_cache[email_key] = _email_exists(email_key)
            return email_exists_cache[email_key]

        def username_exists_cached(username_key):
            if username_key not in username_exists_cache:
                username_exists_cache[username_key] = _username_exists(username_key)
            return username_exists_cache[username_key]

        for row_index, row in enumerate(rows[1:], start=2):
            if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            total_rows += 1

            name = get_value(row, ['name', 'full_name', 'fullname'])
            email = get_value(row, ['email', 'mail', 'email_address']).lower()
            raw_username = get_value(row, ['username', 'user_name', 'user'])
            raw_whatsapp_number = get_value(row, ['whatsapp_number', 'whatsapp', 'phone', 'mobile', 'whatsappnumber'])

            if not email or not raw_username or not raw_whatsapp_number:
                missing_required_rows += 1
                if len(missing_required_details) < max_preview_details:
                    missing_fields = []
                    if not email:
                        missing_fields.append('email')
                    if not raw_username:
                        missing_fields.append('username')
                    if not raw_whatsapp_number:
                        missing_fields.append('whatsapp_number')
                    missing_required_details.append({
                        'row': row_index,
                        'missing_fields': missing_fields,
                    })
                skipped_rows += 1
                continue

            if not is_valid_email(email):
                invalid_email_rows += 1
                if len(invalid_email_details) < max_preview_details:
                    invalid_email_details.append({
                        'row': row_index,
                        'email': email,
                    })
                skipped_rows += 1
                continue

            email_key = _email_key(email)
            if email_key in seen_email_keys:
                duplicate_email_rows += 1
                if len(duplicate_email_in_file_details) < max_preview_details:
                    duplicate_email_in_file_details.append({
                        'row': row_index,
                        'email': email,
                    })
            elif email_exists_cached(email_key):
                existing_email_rows += 1
                if len(existing_email_details) < max_preview_details:
                    existing_email_details.append({
                        'row': row_index,
                        'email': email,
                    })

            seen_email_keys.add(email_key)

            whatsapp_number = normalize_indian_whatsapp_number(raw_whatsapp_number)
            if not whatsapp_number:
                invalid_whatsapp_rows += 1
                if len(invalid_whatsapp_details) < max_preview_details:
                    invalid_whatsapp_details.append({
                        'row': row_index,
                        'whatsapp_number': str(raw_whatsapp_number or ''),
                    })
                skipped_rows += 1
                continue

            username_key = _username_key(raw_username)
            if username_key in seen_username_keys:
                duplicate_username_rows += 1
                if len(duplicate_username_in_file_details) < max_preview_details:
                    duplicate_username_in_file_details.append({
                        'row': row_index,
                        'username': str(raw_username or ''),
                    })
                skipped_rows += 1
                continue

            if username_exists_cached(username_key):
                existing_username_rows += 1
                if len(existing_username_details) < max_preview_details:
                    existing_username_details.append({
                        'row': row_index,
                        'username': str(raw_username or ''),
                    })
                skipped_rows += 1
                continue

            seen_username_keys.add(username_key)

            valid_rows += 1

        return jsonify({
            'success': True,
            'message': 'Preview generated successfully',
            'preview': {
                'total_rows': total_rows,
                'valid_rows': valid_rows,
                'skipped_rows': skipped_rows,
                'missing_required_rows': missing_required_rows,
                'duplicate_email_rows': duplicate_email_rows,
                'existing_email_rows': existing_email_rows,
                'email_conflict_rows': duplicate_email_rows + existing_email_rows,
                'duplicate_username_rows': duplicate_username_rows,
                'existing_username_rows': existing_username_rows,
                'username_conflict_rows': duplicate_username_rows + existing_username_rows,
                'invalid_email_rows': invalid_email_rows,
                'invalid_whatsapp_rows': invalid_whatsapp_rows,
                'detected_headers': [key for key in headers if key],
                'required_header_status': {
                    'email': bool(
                        header_map.get('email') is not None
                        or header_map.get('mail') is not None
                        or header_map.get('email_address') is not None
                    ),
                    'username': bool(
                        header_map.get('username') is not None
                        or header_map.get('user_name') is not None
                        or header_map.get('user') is not None
                    ),
                    'whatsapp_number': bool(
                        header_map.get('whatsapp_number') is not None
                        or header_map.get('whatsapp') is not None
                        or header_map.get('phone') is not None
                        or header_map.get('mobile') is not None
                        or header_map.get('whatsappnumber') is not None
                    ),
                },
                'issue_details': {
                    'missing_required': missing_required_details,
                    'duplicate_email_in_file': duplicate_email_in_file_details,
                    'duplicate_email_existing': existing_email_details,
                    'invalid_email': invalid_email_details,
                    'invalid_whatsapp': invalid_whatsapp_details,
                    'duplicate_username_in_file': duplicate_username_in_file_details,
                    'duplicate_username_existing': existing_username_details,
                    'truncated_to': max_preview_details,
                },
            }
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Preview failed: {str(e)}'
        }), 400

@admin_bp.route('/import/excel', methods=['POST'])
@token_required
@admin_required
def import_users_from_excel(user):
    """Import users from Excel file"""
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    
    file = request.files['file']
    event_id = request.form.get('event_id')  # Get selected event/chatbot ID
    default_role = str(request.form.get('default_role', 'user') or 'user').strip().lower()

    if not str(event_id or '').strip():
        return jsonify({'success': False, 'message': 'Please select an event before importing users'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': 'Only .xlsx files are allowed'}), 400

    if default_role not in VALID_USER_ROLES:
        default_role = 'user'
    
    # Validate event_id
    chatbot = None
    try:
        chatbot = Chatbot.query.get(int(event_id))
        if not chatbot:
            return jsonify({'success': False, 'message': 'Selected event/chatbot not found'}), 404
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Invalid event ID'}), 400
    
    try:
        workbook = load_workbook(filename=BytesIO(file.read()), data_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({'success': False, 'message': 'Excel file has no data rows'}), 400

        headers = [_normalize_excel_header_key(cell) if cell is not None else '' for cell in rows[0]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        def get_value(row, keys, default=''):
            for key in keys:
                index = header_map.get(key)
                if index is not None and index < len(row):
                    value = row[index]
                    if value is not None:
                        return _stringify_excel_cell(value)
            return default

        def parse_bool(value, default=True):
            if value is None or value == '':
                return default
            return str(value).strip().lower() in ['1', 'true', 'yes', 'y', 'active']

        credentials = []
        skipped = 0
        total_rows = 0
        missing_required_rows = 0
        invalid_email_rows = 0
        invalid_whatsapp_rows = 0
        parsed_rows = []
        seen_username_keys = set()
        reserved_email_keys = set()
        auto_adjusted_email_rows = 0
        duplicate_usernames_in_file = []

        for row_index, row in enumerate(rows[1:], start=2):
            if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            total_rows += 1

            name = get_value(row, ['name', 'full_name', 'fullname'])
            email = get_value(row, ['email', 'mail', 'email_address']).lower()
            role = get_value(row, ['role', 'user_role'], default_role).lower()
            raw_username = get_value(row, ['username', 'user_name', 'user'])
            raw_whatsapp_number = get_value(row, ['whatsapp_number', 'whatsapp', 'phone', 'mobile', 'whatsappnumber'])
            password = get_value(row, ['password', 'pass'])
            active_value = get_value(row, ['active', 'is_active'], '')

            if not email or not raw_username or not raw_whatsapp_number:
                missing_required_rows += 1
                skipped += 1
                continue

            if not is_valid_email(email):
                invalid_email_rows += 1
                skipped += 1
                continue

            effective_email, email_auto_adjusted = _make_unique_email_with_reserved(
                email,
                reserved_email_keys,
            )
            if email_auto_adjusted:
                auto_adjusted_email_rows += 1

            reserved_email_keys.add(_email_key(effective_email))

            normalized_whatsapp_number = normalize_indian_whatsapp_number(raw_whatsapp_number)
            if not normalized_whatsapp_number:
                invalid_whatsapp_rows += 1
                skipped += 1
                continue

            if role not in VALID_USER_ROLES:
                role = default_role

            # Prevent regular admin users from importing/creating admin accounts
            if role == 'admin' and getattr(user, 'role', None) == 'admin':
                role = 'user'

            username = str(raw_username).strip()
            username_key = _username_key(username)
            if username_key in seen_username_keys:
                duplicate_usernames_in_file.append({
                    'row': row_index,
                    'username': username,
                })
                skipped += 1
                continue

            seen_username_keys.add(username_key)
            display_name = (name or username or email).strip()
            if not password:
                password = '123'

            parsed_rows.append({
                'row_index': row_index,
                'name': display_name,
                'email': effective_email,
                'username': username,
                'username_key': username_key,
                'whatsapp_number': normalized_whatsapp_number,
                'role': role,
                'active': parse_bool(active_value, True),
                'password': password,
            })

        existing_username_keys = _existing_username_keys({item['username_key'] for item in parsed_rows})
        existing_username_conflicts = []
        valid_rows_to_import = []

        for parsed in parsed_rows:
            if parsed['username_key'] in existing_username_keys:
                existing_username_conflicts.append({
                    'row': parsed['row_index'],
                    'username': parsed['username'],
                })
                skipped += 1
                continue
            valid_rows_to_import.append(parsed)

        if duplicate_usernames_in_file or existing_username_conflicts:
            return jsonify({
                'success': False,
                'message': 'Import blocked: duplicate usernames found. Fix duplicate usernames before importing.',
                'duplicate_username_rows': {
                    'in_file': duplicate_usernames_in_file,
                    'existing_in_system': existing_username_conflicts,
                },
                'duplicate_username_count': len(duplicate_usernames_in_file) + len(existing_username_conflicts),
                'import_summary': {
                    'total_rows': total_rows,
                    'ready_rows': len(valid_rows_to_import),
                    'skipped_rows': skipped,
                    'missing_required_rows': missing_required_rows,
                    'invalid_email_rows': invalid_email_rows,
                    'invalid_whatsapp_rows': invalid_whatsapp_rows,
                    'auto_adjusted_email_rows': auto_adjusted_email_rows,
                },
                'skipped': skipped,
            }), 409

        for parsed in valid_rows_to_import:
            new_user = User(
                name=parsed['name'],
                email=parsed['email'],
                username=parsed['username'],
                whatsapp_number=parsed['whatsapp_number'],
                role=parsed['role'],
                active=parsed['active']
            )
            new_user.set_password(parsed['password'])
            db.session.add(new_user)
            db.session.flush()  # Get the user ID
            
            # Associate user with the selected chatbot/event
            if chatbot:
                existing_participant = ChatbotParticipant.query.filter_by(
                    chatbot_id=chatbot.id,
                    user_id=new_user.id
                ).first()
                
                if not existing_participant:
                    participant = ChatbotParticipant(
                        chatbot_id=chatbot.id,
                        user_id=new_user.id
                    )
                    db.session.add(participant)


            credentials.append({
                'name': parsed['name'],
                'role': parsed['role'],
                'username': parsed['username'],
                'password': parsed['password'],
                'email': parsed['email'],
                'whatsapp_number': parsed['whatsapp_number'],
                'allowed_events': [chatbot.event_name] if chatbot and chatbot.event_name else [],
                'allowed_chatbots': [chatbot.name] if chatbot and chatbot.name else []
            })

        if not credentials:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': 'No valid users imported. Check required columns and data format.',
                'import_summary': {
                    'total_rows': total_rows,
                    'ready_rows': 0,
                    'skipped_rows': skipped,
                    'missing_required_rows': missing_required_rows,
                    'invalid_email_rows': invalid_email_rows,
                    'invalid_whatsapp_rows': invalid_whatsapp_rows,
                    'auto_adjusted_email_rows': auto_adjusted_email_rows,
                    'duplicate_username_in_file_rows': len(duplicate_usernames_in_file),
                    'duplicate_username_existing_rows': len(existing_username_conflicts),
                }
            }), 400

        db.session.commit()

        event_name = chatbot.event_name if chatbot else 'None'
        # Notification for admin import removed - bell feature is no longer used
        db.session.commit()

        # Send all emails in background thread so admin doesn't wait
        app = current_app._get_current_object()
        thread = threading.Thread(
            target=send_bulk_emails_background,
            args=(app, credentials)
        )
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'message': f'Users imported successfully and assigned to event "{event_name}"! Credentials emails are being sent to {len(credentials)} users.',
            'count': len(credentials),
            'skipped': skipped,
            'auto_adjusted_email_rows': auto_adjusted_email_rows,
            'event_name': event_name,
            'credentials': credentials
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Import failed: {str(e)}'
        }), 400
