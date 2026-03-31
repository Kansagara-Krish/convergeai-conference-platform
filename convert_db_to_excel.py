#!/usr/bin/env python3
"""
Convert PostgreSQL database dump to Excel file with schema details
"""
import os
import sys
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def read_sql_file(filepath):
    """Read SQL file and return content"""
    try:
        with open(filepath, 'rb') as f:
            content = f.read()
        # Try to decode as text, skip binary parts
        try:
            return content.decode('utf-8', errors='ignore')
        except:
            return content.decode('latin-1', errors='ignore')
    except Exception as e:
        print(f"Error reading file: {e}")
        return ""

def extract_create_table_statements(sql_content):
    """Extract CREATE TABLE statements from SQL"""
    # Pattern to match CREATE TABLE statements
    pattern = r'CREATE TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:public\.)?(\w+)\s*\((.*?)\);'
    matches = re.findall(pattern, sql_content, re.DOTALL | re.IGNORECASE)
    
    tables = {}
    for table_name, columns_def in matches:
        # Parse column definitions
        columns = parse_columns(columns_def)
        tables[table_name] = columns
    
    return tables

def parse_columns(columns_def):
    """Parse column definitions from CREATE TABLE statement"""
    columns = []
    # Split by comma, but be careful with function definitions
    lines = columns_def.split(',')
    
    for line in lines:
        line = line.strip()
        if not line or line.startswith('CONSTRAINT') or line.startswith('PRIMARY KEY') or line.startswith('FOREIGN KEY') or line.startswith('INDEX'):
            continue
        
        # Extract column name and type
        parts = line.split(None, 1)
        if len(parts) >= 2:
            col_name = parts[0].strip('"`')
            col_type = parts[1]
            
            # Clean up column type
            col_type = re.sub(r'\s+', ' ', col_type).strip()
            
            columns.append({
                'name': col_name,
                'type': col_type
            })
    
    return columns

def create_excel_file(tables, output_filepath):
    """Create Excel file with database schema"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet
    summary_sheet = wb.create_sheet('Summary', 0)
    summary_sheet['A1'] = 'Database Schema Overview'
    summary_sheet['A1'].font = Font(size=14, bold=True)
    summary_sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    summary_sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    
    summary_sheet['A3'] = 'Table Name'
    summary_sheet['B3'] = 'Column Count'
    
    # Style header
    for col in ['A', 'B']:
        summary_sheet[f'{col}3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        summary_sheet[f'{col}3'].font = Font(bold=True)
    
    # Add summary data
    row = 4
    for table_name, columns in sorted(tables.items()):
        summary_sheet[f'A{row}'] = table_name
        summary_sheet[f'B{row}'] = len(columns)
        row += 1
    
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    
    # Create detailed sheet for each table
    for table_name, columns in sorted(tables.items()):
        sheet = wb.create_sheet(table_name[:31])  # Excel sheet name limit is 31
        
        # Header
        sheet['A1'] = f'Table: {table_name}'
        sheet['A1'].font = Font(size=12, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sheet.merge_cells('A1:C1')
        
        # Column headers
        headers = ['Column Name', 'Data Type', 'Details']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add column data
        row = 4
        for column in columns:
            sheet[f'A{row}'] = column['name']
            sheet[f'B{row}'] = column['type']
            
            # Extract constraints/details
            details = []
            col_type_lower = column['type'].lower()
            if 'primary key' in col_type_lower:
                details.append('PK')
            if 'not null' in col_type_lower:
                details.append('NOT NULL')
            if 'unique' in col_type_lower:
                details.append('UNIQUE')
            if 'default' in col_type_lower:
                details.append('HAS DEFAULT')
            if 'references' in col_type_lower or 'foreign key' in col_type_lower:
                details.append('FK')
            
            sheet[f'C{row}'] = ', '.join(details)
            row += 1
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 30
    
    # Save workbook
    wb.save(output_filepath)
    print(f"✓ Excel file created successfully: {output_filepath}")
    print(f"✓ Total tables: {len(tables)}")

def main():
    # Get the SQL file path
    sql_file = 'convergeai_conference_db.sql'
    current_dir = Path(__file__).parent
    sql_filepath = current_dir / sql_file
    
    if not sql_filepath.exists():
        print(f"Error: {sql_file} not found in {current_dir}")
        sys.exit(1)
    
    print(f"📖 Reading SQL file: {sql_filepath}")
    sql_content = read_sql_file(str(sql_filepath))
    
    print("🔍 Extracting table definitions...")
    tables = extract_create_table_statements(sql_content)
    
    if not tables:
        print("⚠ No tables found. The SQL file might be binary or in an unsupported format.")
        print("Attempting alternative parsing method...")
        
        # Try to find any reference information in the SQL
        print("Hint: This appears to be a PostgreSQL binary dump. Consider using:")
        print(f"  pg_restore --list {sql_file}")
        print(f"Or convert it first: pg_restore -d mydb {sql_file}")
        sys.exit(1)
    
    # Create output file
    output_file = current_dir / 'database_schema.xlsx'
    print(f"📝 Creating Excel file: {output_file}")
    create_excel_file(tables, str(output_file))
    print("\n✅ Conversion completed successfully!")

if __name__ == '__main__':
    main()
